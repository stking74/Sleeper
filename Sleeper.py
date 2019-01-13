# -*- coding: utf-8 -*-

import esipy as api
import datetime
import time
import numpy as np
import pickle
import os

def crawldir(topdir=[], ext='sxm'):
    '''
    Crawls through given directory topdir, returns list of all files with 
    extension matching ext
    '''
    import re
    fn = dict()
    for root, dirs, files in os.walk(topdir):
              for name in files:
              
                if len(re.findall('\.'+ext,name)):
                    addname = os.path.join(root,name)

                    if root in fn.keys():
                        fn[root].append(addname)

                    else:
                        fn[root] = [addname]    
    return fn

class Sleeper():
    
    def __init__(self):
        print('Initiailizing app...')
        self.esi_app = api.EsiApp()
        print('Updating swagger...')
        self.app = self.esi_app.get_latest_swagger
        print('Initializing client...')
        self.client = api.EsiClient(retry_requests=True, 
                                    headers={'User-Agent':'Sleeper \\ <0.3.0>'},
                                    raw_body_only=False)
        self.root_dir = os.getcwd()
        self.store_dir = os.path.join(self.root_dir, 'data_dumps')
        self.settings_fname = 'sleeper_settings.sl'
        return
    
    def _update_region_list(self):
        
        print('Refreshing region metadata...')
        operation = self.app.op['get_universe_regions']()
        response = self.client.request(operation)
        region_ids = response.data
        self.region_list = {}
        i = 1
        for region_id in region_ids:
#            print(f'{i}/{len(region_ids)}')
            operation = self.app.op['get_universe_regions_region_id'](
                region_id=region_id)
            response = self.client.request(operation)
            region_name = response.data['name']
            self.region_list[region_name] = dict(response.data)
            for name, data in self.region_list.items():
                self.region_list[name]['constellations'] = list(self.region_list[name]['constellations'])
            i += 1
        return self.region_list
    
    def market_dump(self):
        
        def _request_region_market_orders(region_id=10000002, type_id=34, order_type='all'):
            page = 1
            while True:
                operation = self.app.op['get_markets_region_id_orders'](region_id=region_id, order_type=order_type, page=page)
                response = self.client.request(operation)
                pull_time = datetime.datetime.now()
                time.sleep(0.25)
                if page == 1:
                    orders = [dict(entry) for entry in response.data]
                    for order in orders:
                        order['timestamps'] = pull_time
                    if len(orders)==0:
                        break
                    page += 1
                    continue
                else:
                    new_orders = [dict(entry) for entry in response.data]
                    for order in new_orders:
                        order['timestamps'] = pull_time
                    if len(new_orders) == 0:
                        break
                    for entry in new_orders:
                        orders.append(entry)
                page += 1
            return orders
        
        orders = {}
        print('Scraping market data')
        print('Region:')
        for region in self.region_list.items():
            
            name, region_data = region
            print(name)
            region_id = region_data['region_id']
            orders[name] = _request_region_market_orders(region_id, order_type='all')
        refresh_time = datetime.datetime.now()
        filename_timestamp = refresh_time.strftime('%c')
        os.chdir(self.store_dir)
        pik_filename = 'market_dump-'+str(refresh_time)[:10]+'.pik'
        pik_file = open(pik_filename, 'wb')
        pickle.dump(orders, pik_file)
        pik_file.close()
        os.chdir(self.root_dir)
        
        return
    
    def _load_dumpfile_(self, fname):
        with open(fname, 'rb') as f:
            raw_catalog = pickle.load(f)  
        print('Cataloging orders...')
        catalog = []
        for key, region in raw_catalog.items():
            for order in region:
                duration = order['duration']
                is_buy_order = order['is_buy_order']
                issued = order['issued']
                location_id = order['location_id']
                min_volume = order['min_volume']
                order_id = order['order_id']
                price = order['price']
                _range = order['range']
                system_id = order['system_id']
                timestamps = order['timestamps']
                type_id = order['type_id']
                volume_remain = int(order['volume_remain'])
                volume_total = order['volume_total']
                catalog.append(Order(duration, is_buy_order, issued, location_id, min_volume,
                             order_id, price, _range, system_id, timestamps, type_id,
                             volume_remain, volume_total))
        return catalog
    
    
    def aggregate_data(self, data_directory):
        print('Aggregating order data...')
        data_directory_contents = os.listdir(data_directory)
        split_contents = [item.split('.') for item in data_directory_contents]
        target_file_list = []
        for idx, item in enumerate(split_contents):
            if item[1] == 'pik':
                target_file_list.append(data_directory_contents[idx])
        aggregated_raw_data = {}
        for file in target_file_list:
            pull_date = file[12:22]
            file = os.path.join(data_directory, file)
            with open(file, 'rb') as pik_file:
                pickle_data = pickle.load(pik_file)
            aggregated_raw_data[pull_date] = pickle_data
        
        print('Extracting order IDs...')
        order_catalog = {}
        for date in aggregated_raw_data.values():
            for region in date.values():
                for order in region:
                    order_id = str(order['order_id'])
                    if order_id in order_catalog.keys():
                        order_catalog[order_id].append(order)
                    else:
                        order_catalog[order_id] = [order]
        '''
        order_ids = list(set(order_ids))

        orders = {}
        for order_id in order_ids:
            new_key = str(order_id)
            orders[new_key] = {}
            
        print('Cataloging orders...')
        for idx, order_id in enumerate(orders.keys()):
            for date in aggregated_raw_data.values():
                for region in date.values():
                    for order in region:
                        #!!! This if condition below is not working as intended
                        if order['order_id'] == order_id:
                            orders[order_id].append(order)
            if idx % 1000 == 0:
                print(int(idx),'/',len(orders))
        '''
        return order_catalog
    
    def _load_settings_file_(self):
        dir_dump = crawldir(self.root_dir, 'sl')
        if len(dir_dump) == 0:
            self._new_settings_file_()
        return
    
    def _new_settings_file_():
        return
    
    def _aggregate_weekly_(self):
        week_ago = datetime.datetime.now() - datetime.timedelta(days=7)
        weekly_dumps = []
        os.chdir(self.store_dir)
        contents = os.listdir(self.store_dir)
        files = []
        for item in contents:
            if os.path.isfile(item):
                fname, ext = item.split('.')
                if ext == 'pik':
                    weekly_dumps.append(item)
        region_list = self._update_region_list()
        catalog = {key:[] for key in region_list.keys()}
        order_ids = set()
        order_id_list = {key:[] for key in region_list.keys()}
        previous_length = 0
        for file in weekly_dumps:
            f_timestamp = file[12:22]
            timestamp = datetime.datetime.strptime(f_timestamp, '%Y-%m-%d')
            if timestamp > week_ago:
                print('loading file:', file)
                with open(file, 'rb') as f:
                    data_dump = pickle.load(f)
                for key, region in data_dump.items():
                    region_catalog = catalog[key]
                    ti = time.time()
                    oids = [order['order_id'] for order in region]
                    for order in region:
                        oid = order['order_id']
                        if oid not in order_ids:
                            order_ids.add(oid)
                            order_id_list[key].append(oid)
                            order['volume_remain'] = [order['volume_remain']]
                            try:
                                order['timestamps'] = [order['timestamps']]
                            except KeyError:
                                order['timestamps'] = [timestamp]
                            order['price'] = [order['price']]
                            region_catalog.append(order)
                        else:
                            i = order_id_list[key].index(oid)
                            check = region_catalog[i]
                            cid = check['order_id']
                            if cid == oid:
                                check['volume_remain'].append(order['volume_remain'])
                                check['price'].append(order['price'])
                                try:
                                    check['timestamps'].append(order['timestamps'])
                                except KeyError:
                                    check['timestamps'].append(timestamp)
                                region_catalog[i] = check
#                               print(oid, len(order_ids), te)
                                continue
                    te = time.time() - ti
                    print('%s: %i %f'%(key, len(region), te))
#                        print(oid, len(order_ids), te)
                running_length = sum(len(region) for region in catalog)
                print('Current catalog length:', running_length)
                new_orders = running_length - previous_length
                previous_length = running_length
                print('Added',new_orders,'new orders.')
        return catalog

class Order:
    '''
    Object for handling and organization of order data. Attributes correspond
    to values available from Swagger through regional market order requests. 
    '''
    def __init__(self, duration, is_buy_order, issued, location_id, min_volume,
                 order_id, price, _range, system_id, timestamps, type_id,
                 volume_remain, volume_total):
        '''
        Initialize Order, assign attributes from individual order entries in 
        raw data dump.
        
        Input:
        --------
            duration : int
                Duration of order, in hours or days
            is_buy_order : bool
                True if buy order, False if sell order
            issued : datetime.datetime
                Timestamp for order creation
            location_id : int
                Unique location identifier for order
            min_volume : int
                Minimum volume allowed per transaction
            order_id : int
                Unique order identification number
            price : list of floats
                Order unit price, in the order in which they were captured in a market dump
            _range : str
                Distance from location within which transaction may occur
            system_id : int
                Unique identifier of system in which order was placed, redundant information also provided indirectly by location_id
            timestamps : list of datetime.datetime objects
                Timestamps for each market dump in which order was located
            type_id : int
                Unique identifier for type of item being sold/bought in order
            volume_remain : int
                Total number of units still available/desired in order
            volume_total : int
                Total number of units available/desired at order creation
        '''
        self.duration = duration
        self.is_buy_order = is_buy_order
        self.issued = issued
        self.location_id = location_id
        self.min_volume = min_volume
        self.order_id = order_id
        self.price = price
        self._range = _range
        self.system_id = system_id
        self.timestamps = timestamps
        self.type_id = type_id
        self.volume_remain = volume_remain
        self.volume_total = volume_total
        return
    
    def mean_price(self):
        '''
        
        Calculates the average of all prices identified for this order.
        '''
        prices = [price for price in self.price]
        average_price = np.mean(prices)
        return average_price
    
    def delta_price(self):
        '''
        Calculates total change in listed price since order creation.
        '''
        prices = [price for price in self.price]
        if len(prices) < 2:
            price_change = None
        else:
            price_change = prices[-1] - prices[0]
        return price_change
    
    def age(self):
        '''
        Calculates total age of order since issuance.
        '''
        now = datetime.datetime.now()
        order_age = now - self.issued
        return order_age
    
class Reporter:
    '''
    Object for organization and handling of Sleeper data for the purposes
    of compilation of aggregate data into a meaningful format. 
    '''
    
    def __init__(self, data_dir, resource_dir, report_dir):
        self.data_dir = data_dir
        self.resource_dir = resource_dir
        self.report_dir = report_dir
        return
    
    def list2sheet(self, list_data, worksheet):
        '''
        Converts a list of lists into a formatted openpyxl worksheet, with each nested
        list representing a separate row of the final worksheet
        '''
        for row, data in enumerate(list_data):
            row += 1
            for column, value in enumerate(data):
                column += 1
                worksheet.cell(row=row, column=column, value=value)
        return worksheet
    
    def generate_daily_report(self, date=None):
        import openpyxl as op
        import scipy.stats as stats
        import matplotlib.pyplot as plt
        import datetime
        
        #Data read-in and organization, and analysis setup
        today = datetime.datetime.today().date()
        yesterday = today - datetime.timedelta(days=1)
        no_previous_data = False
        valid_files = []
        # Identify most recently produced data dump
        for root, dirs, files in os.walk(self.data_dir):
            if root == self.data_dir:
                for filename in files:
                    try:
                        f_timestamp = filename[12:22]
                        timestamp = datetime.datetime.strptime(f_timestamp, '%Y-%m-%d').date()
                        valid_files.append((filename, timestamp))
                    except:
                        continue
        if len(valid_files) > 1:
            most_recent = valid_files[0]
            for f in valid_files[1:]:
                if f[1] > most_recent[1]:
                    most_recent = f
        else:
            most_recent = valid_files[0]
        most_recent_fname = most_recent[0]
        ##### Keep everything below this line in a state where it can be copy-pasted into retroactive report generator script
        timestamp = most_recent[1]
        newdir_name = 'sleeper_daily_'+most_recent_fname[12:22]
        report_dir = os.path.join(self.report_dir, newdir_name)
        try:
            os.mkdir(report_dir)
        except:
            pass
        figure_dir = os.path.join(report_dir,'figures')
        try:
            os.mkdir(figure_dir)
        except:
            pass
        print('Loading file: %s'%most_recent_fname)
        most_recent_fname = os.path.join(self.data_dir, most_recent_fname)
        with open(most_recent_fname, 'rb') as f:
            raw_catalog = pickle.load(f)  
        print('Cataloging orders...')
        catalog = []
        for key, region in raw_catalog.items():
            for order in region:
                duration = order['duration']
                is_buy_order = order['is_buy_order']
                issued = order['issued']
                location_id = order['location_id']
                min_volume = order['min_volume']
                order_id = order['order_id']
                price = order['price']
                _range = order['range']
                system_id = order['system_id']
                timestamps = order['timestamps']
                type_id = order['type_id']
                volume_remain = int(order['volume_remain'])
                volume_total = order['volume_total']
                catalog.append(Order(duration, is_buy_order, issued, location_id, min_volume,
                             order_id, price, _range, system_id, timestamps, type_id,
                             volume_remain, volume_total))
        type_ids = np.array([order.type_id for order in catalog])
        duration_bins = list(set([order.duration for order in catalog]))
        
            
        '''
        Ores
        '''
        print('Aggregating ore data...')
        report = {}
        report['Ores'] = {}
        ore_ids_fname = os.path.join(self.resource_dir, 'ore_ids.txt')
        with open(ore_ids_fname, 'r') as f:
            ore_type_ids = {}
            for line in f:
                tid, name, extra = line.split(',')
                ore_type_ids[name] = int(tid)
        ice_ids_fname = os.path.join(self.resource_dir, 'ice_ids.txt')
        with open(ice_ids_fname, 'r') as f:
            ice_type_ids = {}
            for line in f:
                tid, name = line.split(',')
                if '\n' in name:
                    name = name[:-1]
                ice_type_ids[name] = int(tid)
        ore_groups = {'Arkonor':{'Uncompressed':['Arkonor','Crimson Arkonor', 'Prime Arkonor', 'Flawless Arkonor']},
                      'Bistot':{'Uncompressed':['Bistot', 'Triclinic Bistot', 'Monoclinic Bistot', 'Cubic Bistot']},
                      'Crokite':{'Uncompressed':['Crokite', 'Sharp Crokite', 'Crystalline Crokite', 'Pellucid Crokite']},
                      'Ochre':{'Uncompressed':['Dark Ochre', 'Onyx Ochre', 'Obsidian Ochre', 'Jet Ochre']},
                      'Gneiss':{'Uncompressed':['Gneiss', 'Iridescent Gneiss', 'Prismatic Gneiss', 'Brilliant Gneiss']},
                      'Hedbergite':{'Uncompressed':['Hedbergite', 'Vitric Hedbergite', 'Glazed Hedbergite', 'Lustrous Hedbergite']},
                      'Hemorphite':{'Uncompressed':['Hemorphite','Vivid Hemorphite', 'Radiant Hemorphite', 'Scintillating Hemorphite']},
                      'Jaspet':{'Uncompressed':['Jaspet', 'Pure Jaspet', 'Pristine Jaspet', 'Immaculate Jaspet']},
                      'Kernite':{'Uncompressed':['Kernite', 'Luminous Kernite', 'Fiery Kernite', 'Resplendant Kernite']},
                      'Mercoxit':{'Uncompressed':['Mercoxit', 'Magma Mercoxit', 'Vitreous Mercoxit']},
                      'Omber':{'Uncompressed':['Omber', 'Silvery Omber','Golden Omber','Platinoid Omber']},
                      'Plagioclase':{'Uncompressed':['Plagioclase', 'Azure Plagioclase', 'Rich Plagioclase', 'Sparkling Plagioclase']},
                      'Pyroxeres':{'Uncompressed':['Pyroxeres', 'Solid Pyroxeres', 'Viscous Pyroxeres', 'Opulent Pyroxeres']},
                      'Scordite':{'Uncompressed':['Scordite','Condensed Scordite', 'Massive Scordite', 'Glossy Scordite']},
                      'Spodumain':{'Uncompressed':['Spodumain', 'Bright Spodumain', 'Gleaming Spodumain', 'Dazzling Spodumain']},
                      'Veldspar':{'Uncompressed':['Veldspar', 'Concentrated Veldspar', 'Dense Veldspar', 'Stable Veldspar']}}
        max_sec = {'Veldspar':1.0,'Scordite':1.0,'Pyroxeres':0.9,'Plagioclase':0.8,
                   'Omber':0.7,'Kernite':0.6,'Jaspet':0.4,'Hemorphite':0.2,'Hedbergite':0.2,
                   'Gneiss':-0.4,'Ochre':-0.2,'Spodumain':-0.5,'Crokite':-0.5,'Bistot':-0.6,
                   'Arkonor':-0.7,'Mercoxit':-0.8}
        for ore_group in ore_groups.values():
            ore_group['Compressed'] = ['Compressed ' + string for string in ore_group['Uncompressed']]
        min_rstd = 100000000000000000000
        max_rstd = 0
        for ore_type, ore_type_id in ore_type_ids.items():
            if ore_type.split(' ')[0] == 'Compressed':
                compressed_bool = True
            else:
                compressed_bool = False
            ore_report = {}
            ore_report['figures'] = {}
            matching_order_indices = np.where(type_ids == ore_type_id)[0]
            ore_orders = [catalog[i] for i in matching_order_indices]
            current_prices = [order.price for order in ore_orders]
            buy_sell = [order.is_buy_order for order in ore_orders]
            ore_report['n_orders'] = len(ore_orders)
            buy_order_indices = np.where(np.array(buy_sell)==True)[0]
            ore_buy_prices = [ore_orders[i].price for i in buy_order_indices]
            sell_order_indices = np.where(np.array(buy_sell)==False)[0]
            ore_sell_prices = [ore_orders[i].price for i in sell_order_indices]
            ore_report['percent_buy'] = len(np.where(np.array(buy_sell)==True)[0])/len(ore_orders)*100
            ore_report['n_buy_orders'] = len(buy_order_indices)
            ore_report['n_sell_orders']  = len(sell_order_indices)
            ore_report['mean_buy_price'] = np.mean(ore_buy_prices)
            ore_report['std_buy_price'] = np.std(ore_buy_prices)
            ore_report['max_buy_price'] = np.max(ore_buy_prices)
            ore_report['min_buy_price'] = np.min(ore_buy_prices)
            ore_report['total_buy_volume'] = 0
            for i in buy_order_indices:
                order = ore_orders[i]
                volume = int(order.volume_remain)
                ore_report['total_buy_volume'] += volume
            ore_report['gross_estimated_buy_value'] = ore_report['mean_buy_price'] * ore_report['total_buy_volume']
            ore_report['mean_sell_price'] = np.mean(ore_sell_prices)
            ore_report['std_sell_price'] = np.std(ore_sell_prices)
            ore_report['max_sell_price'] = np.max(ore_sell_prices)
            ore_report['min_sell_price'] = np.min(ore_sell_prices)
            ore_report['total_sell_volume'] = 0
            for order in ore_orders:
                order = ore_orders[i]
                volume = int(order.volume_remain)
                ore_report['total_sell_volume'] += volume
            ore_report['gross_estimated_sell_value'] = ore_report['mean_sell_price'] * ore_report['total_sell_volume']
            ore_report['global_current_prices'] = np.mean(current_prices)
            ore_report['global_price_std'] = np.std(current_prices)
            ore_report['global_price_rstd'] = (np.std(current_prices) / np.mean(current_prices)) * 100
            if ore_report['global_price_rstd'] > max_rstd:
                max_rstd = ore_report['global_price_rstd']
                most_unstable_ore = (ore_type, max_rstd)
            if ore_report['global_price_rstd'] < min_rstd:
                min_rstd = ore_report['global_price_rstd']
                most_stable_ore = (ore_type, min_rstd)
            duration_counts = np.zeros_like(duration_bins)
            for order in ore_orders:
                d = order.duration
                duration_counts[duration_bins.index(d)] += 1
            ore_report['figures']['duration_histogram'] = {str(duration_bins[i]):duration_counts[i] for i in range(len(duration_bins))}
            report['Ores'][ore_type] = ore_report
        #Categorical and statistical analyses
        report['Ores']['grouped_volume'] = {}
        report['Ores']['grouped_n_orders'] = {}
        for key, ore_group in ore_groups.items():
            group_n_orders = 0
            group_volume = 0
            for ore_type in ore_group['Uncompressed']:
                tid = ore_type_ids[ore_type]
                matching_order_indices = np.where(type_ids == tid)[0]
                matching_orders = [catalog[i] for i in matching_order_indices]
                group_n_orders += len(matching_orders)
                for order in matching_orders:
                    group_volume += int(order.volume_remain)
            for ore_type in ore_group['Compressed']:
                tid = ore_type_ids[ore_type]
                matching_order_indices = np.where(type_ids == tid)[0]
                matching_orders = [catalog[i] for i in matching_order_indices]
                group_n_orders += len(matching_orders)
                for order in matching_orders:
                    group_volume += int(order.volume_remain)
            report['Ores']['grouped_volume'][key] = group_volume
            report['Ores']['grouped_n_orders'][key] = group_n_orders
        report['Ores']['most/least stable'] = (most_stable_ore, most_unstable_ore)
        #Generate figures
        try:
            os.mkdir(os.path.join(figure_dir,'ores'))
        except:
            pass
        report['Ores']['figures'] = {}
        temp = report['Ores']['grouped_n_orders'].items()
        x = [item[0] for item in temp]
        y = [item[1] for item in temp]
        report['Ores']['figures']['grouped_order_count'] = plt.figure()
        plt.pie(y)
        plt.legend(x)
        plt.title('Ore Grouped Order Count')
        figname = os.path.join(os.path.join(figure_dir,'ores'),'Grouped Order Count.png')
        plt.savefig(figname)
        plt.close()
        temp = report['Ores']['grouped_volume'].items()
        x = [item[0] for item in temp]
        y = [item[1] for item in temp]
        report['Ores']['figures']['grouped_unit_volume'] = plt.figure()
        plt.pie(y)
        plt.legend(x)
        plt.title('Ore Grouped Unit Volume')
        figname = os.path.join(os.path.join(figure_dir,'ores'),'Grouped Unit Volume.png')
        plt.savefig(figname)
        plt.close()
        ##Ice subreport
        ice_report = {}
        min_rstd = 100000000000000000000
        max_rstd = 0
        for ice_type, ice_type_id in ice_type_ids.items():
            ir = ice_report[ice_type] = {}
            ir['figures'] = {}
            matching_order_indices = np.where(type_ids == ice_type_id)[0]
            ice_orders = [catalog[i] for i in matching_order_indices]
            current_prices = [order.price for order in ice_orders]
            buy_sell = [order.is_buy_order for order in ice_orders]
            ir['n_orders'] = len(ice_orders)
            buy_order_indices = np.where(np.array(buy_sell)==True)[0]
            ice_buy_prices = [ice_orders[i].price for i in buy_order_indices]
            sell_order_indices = np.where(np.array(buy_sell)==False)[0]
            ice_sell_prices = [ice_orders[i].price for i in sell_order_indices]
            ir['n_buy_orders'] = len(buy_order_indices)
            ir['n_sell_orders']  = len(sell_order_indices)
            ir['mean_buy_price'] = np.mean(ice_buy_prices)
            ir['std_buy_price'] = np.std(ice_buy_prices)
            ir['max_buy_price'] = np.max(ice_buy_prices)
            ir['min_buy_price'] = np.min(ice_buy_prices)
            ir['mean_sell_price'] = np.mean(ice_sell_prices)
            ir['std_sell_price'] = np.std(ice_sell_prices)
            ir['max_sell_price'] = np.max(ice_sell_prices)
            ir['min_sell_price'] = np.min(ice_sell_prices)
            ir['percent_buy'] = len(np.where(np.array(buy_sell)==True)[0])/len(ice_orders)*100
            ir['global_current_prices'] = np.mean(current_prices)
            ir['global_price_std'] = np.std(current_prices)
            ir['global_price_rstd'] = (np.std(current_prices) / np.mean(current_prices)) * 100
            ir['total_volume'] = np.sum([np.sum(order.volume_remain) for order in ice_orders])
            ir['total_buy_volume'] = np.sum([ice_orders[i].volume_remain for i in buy_order_indices])
            ir['total_sell_volume'] = np.sum([ice_orders[i].volume_remain for i in sell_order_indices])
            if ir['global_price_rstd'] > max_rstd:
                max_rstd = ir['global_price_rstd']
                most_unstable_ice = (ice_type, max_rstd)
            if ir['global_price_rstd'] < min_rstd:
                min_rstd = ir['global_price_rstd']
                most_stable_ice = (ice_type, min_rstd)
            duration_counts = np.zeros_like(duration_bins)
            for order in ice_orders:
                d = order.duration
                duration_counts[duration_bins.index(d)] += 1
            ir['figures']['duration_histogram'] = {str(duration_bins[i]):duration_counts[i] for i in range(len(duration_bins))}
        report['Ices'] = ice_report
        #Write Ore Report
        wb = op.Workbook()
        ore_group_names = [key for key in ore_groups.keys()]
        summary_sheet = wb.active
        ore_data = report['Ores']
        try:
            os.mkdir(os.path.join(figure_dir, 'ores'))
        except:
            pass
        uncompressed_metrics = [key for key in ore_data['Arkonor'].keys()]
        compressed_metrics = [key for key in ore_data['Compressed Arkonor'].keys()]
        for key, ore_group in ore_groups.items():
            lines = []
            ws = wb.create_sheet(key)
            ws.cell(1,1,'Ore')
            ws.cell(2,1,'Uncompressed')
            for column, value in enumerate(uncompressed_metrics):
                if value == 'figures':
                    continue
                column += 2
                ws.cell(2, column, value)
            for row, ore_name in enumerate(ore_group['Uncompressed']):
                ws.cell(row+3, 1, ore_name)
                ore_report = ore_data[ore_name]
                for column, (key, value) in enumerate(ore_report.items()):
                    if key == 'figures':
                        continue
                    column += 2
                    ws.cell(row+3, column, value)
            ws.cell(7,1,'Compressed')
            for column, value in enumerate(uncompressed_metrics):
                if value == 'figures':
                    continue
                column += 2
                ws.cell(7, column, value)
            for row, ore_name in enumerate(ore_group['Compressed']):
                ws.cell(row+8, 1, ore_name)
                ore_report = ore_data[ore_name]
                for column, (key, value) in enumerate(ore_report.items()):
                    if key == 'figures':
                        continue
                    column += 2
                    ws.cell(row+8, column, value)
        # Summary sheet
        summary_sheet.title = 'Summary'
        summary_header = ['','Ore Group', 'Max Sec', 'Group Average Unit Sell Price',
                          'Group RStDev Unit Sell Price','Group Average Unit Buy Price',
                          'Group RStDev Unit Buy Price', 'n buy orders', 'n sell orders',
                          'Sell Volume','Buy Volume','Sell Gross','Buy Gross']
        lines = [['Uncompressed'],
                 summary_header,
                 ['']]
        categorical_column = ['High Sec','Veldspar','Scordite','Pyroxeres','Plagioclase',
                              'Omber','Kernite','','Low','Jaspet','Hemorphite','Hedbergite',
                              '','Null','Gneiss','Ochre','Spodumain','Crokite','Bistot',
                              'Arkonor','Mercoxit']
        min_row = 2
        min_column = 3
        max_column = 2 + len(uncompressed_metrics)
        for ore_group in categorical_column:
            new_row = ['',ore_group]
            if ore_group in ore_group_names:
                new_row.append(max_sec[ore_group])
                group_sheet = wb[ore_group]
                if ore_group == 'Mercoxit': max_row = 5 
                else: max_row = 6
                ore_metrics = {}
                for col in range(min_column, max_column):
                    col_data = [group_sheet.cell(row=row_n, column=col).value for row_n in range(min_row, max_row+1)]
                    header = col_data[0]
                    col_entries = col_data[1:]
                    ore_metrics[header] = col_entries
                group_data = {}
                # Read in base data
                group_data['global_current_prices'] = np.mean(ore_metrics['global_current_prices'])
                group_data['global_price_rstd'] = np.mean(ore_metrics['global_price_rstd'])
                group_data['global_price_std'] = np.mean(ore_metrics['global_price_std'])
                group_data['mean_buy_price'] = np.mean(ore_metrics['mean_buy_price'])
                group_data['std_buy_price'] = np.mean(ore_metrics['std_buy_price'])
                group_data['mean_sell_price'] = np.mean(ore_metrics['mean_sell_price'])
                group_data['std_sell_price'] = np.mean(ore_metrics['std_sell_price'])
                group_data['n_buy_orders'] = np.sum(ore_metrics['n_buy_orders'])
                group_data['n_orders'] = np.sum(ore_metrics['n_orders'])
                group_data['n_sell_orders'] = np.sum(ore_metrics['n_sell_orders'])
                group_data['total_buy_volume'] = np.sum(ore_metrics['total_buy_volume'])
                group_data['total_sell_volume'] = np.sum(ore_metrics['total_sell_volume'])
                # Calculate higher order data
                group_data['total_demand_value'] = group_data['total_buy_volume'] * group_data['mean_buy_price']
                group_data['total_supply_value'] = group_data['total_sell_volume'] * group_data['mean_sell_price']
                group_data['rstd_buy_price'] = group_data['std_buy_price'] / group_data['mean_buy_price']
                group_data['rstd_sell_price'] = group_data['std_sell_price'] / group_data['mean_sell_price']
                # Organize for write to xlsx
                new_row.append(group_data['mean_sell_price'])
                new_row.append(group_data['rstd_sell_price'])
                new_row.append(group_data['mean_buy_price'])
                new_row.append(group_data['rstd_buy_price'])
                new_row.append(group_data['n_buy_orders'])
                new_row.append(group_data['n_sell_orders'])
                new_row.append(group_data['total_sell_volume'])
                new_row.append(group_data['total_buy_volume'])
                new_row.append(group_data['total_supply_value'])
                new_row.append(group_data['total_demand_value'])
            lines.append(new_row)
            
        lines.append(['Compressed'])
        lines.append(summary_header)
        lines.append([''])
        min_row = 7
        min_column = 3
        max_column = 2 + len(compressed_metrics)
        for ore_group in categorical_column:
            new_row = ['',ore_group]
            if ore_group in ore_group_names:
                new_row.append(max_sec[ore_group])
                group_sheet = wb[ore_group]
                if ore_group == 'Mercoxit': 
                    max_row = 10
                else: max_row = 11
                ore_metrics = {}
                for col in range(min_column, max_column):
                    col_data = [group_sheet.cell(row=row_n, column=col).value for row_n in range(min_row, max_row+1)]
                    header = col_data[0]
                    col_entries = col_data[1:]
                    ore_metrics[header] = col_entries
                group_data = {}
                # Read in base data
                group_data['global_current_prices'] = np.mean(ore_metrics['global_current_prices'])
                group_data['global_price_rstd'] = np.mean(ore_metrics['global_price_rstd'])
                group_data['global_price_std'] = np.mean(ore_metrics['global_price_std'])
                group_data['mean_buy_price'] = np.mean(ore_metrics['mean_buy_price'])
                group_data['std_buy_price'] = np.mean(ore_metrics['std_buy_price'])
                group_data['mean_sell_price'] = np.mean(ore_metrics['mean_sell_price'])
                group_data['std_sell_price'] = np.mean(ore_metrics['std_sell_price'])
                group_data['n_buy_orders'] = np.sum(ore_metrics['n_buy_orders'])
                group_data['n_orders'] = np.sum(ore_metrics['n_orders'])
                group_data['n_sell_orders'] = np.sum(ore_metrics['n_sell_orders'])
                group_data['total_buy_volume'] = np.sum(ore_metrics['total_buy_volume'])
                group_data['total_sell_volume'] = np.sum(ore_metrics['total_sell_volume'])
                # Calculate higher order data
                group_data['total_demand_value'] = group_data['total_buy_volume'] * group_data['mean_buy_price']
                group_data['total_supply_value'] = group_data['total_sell_volume'] * group_data['mean_sell_price']
                group_data['rstd_buy_price'] = group_data['std_buy_price'] / group_data['mean_buy_price']
                group_data['rstd_sell_price'] = group_data['std_sell_price'] / group_data['mean_sell_price']
                # Organize for write to xlsx
                new_row.append(group_data['mean_sell_price'])
                new_row.append(group_data['rstd_sell_price'])
                new_row.append(group_data['mean_buy_price'])
                new_row.append(group_data['rstd_buy_price'])
                new_row.append(group_data['n_buy_orders'])
                new_row.append(group_data['n_sell_orders'])
                new_row.append(group_data['total_sell_volume'])
                new_row.append(group_data['total_buy_volume'])
                new_row.append(group_data['total_supply_value'])
                new_row.append(group_data['total_demand_value'])
            lines.append(new_row)
        for idx, line in enumerate(lines):
            idx += 1
            for jdx, entry in enumerate(line):
                jdx += 1
                summary_sheet.cell(idx, jdx, entry)
        ore_report_fname = 'Sleeper_oreReport_daily_'+str(timestamp)+'.xlsx'
        ore_report_fname = os.path.join(report_dir, ore_report_fname)
        wb.save(ore_report_fname)
        #Write Ice Report
        wb = op.Workbook()
        ws = wb.active
        ice_data = report['Ices']
        ice_types = list(ice_data.keys())
        compressed_types = []
        uncompressed_types = []
        for ice in ice_types:
            split_name = ice.split(' ')
            if 'Compressed' in split_name:
                compressed_types.append(ice)
            else:
                uncompressed_types.append(ice)
        try:
            os.mkdir(os.path.join(figure_dir, 'ices'))
        except:
            pass
        uncompressed_metrics = [key for key in ice_data['Blue Ice'].keys()]
        compressed_metrics = [key for key in ice_data['Compressed Blue Ice'].keys()]
        ws.cell(1,1,'Ore')
        ws.cell(2,1,'Uncompressed')
        for column, value in enumerate(uncompressed_metrics):
            if value == 'figures':
                continue
            column += 2
            ws.cell(2, column, value)
        for row, ore_name in enumerate(uncompressed_types):
            ws.cell(row+3, 1, ore_name)
            ore_report = ice_data[ore_name]
            for column, (key, value) in enumerate(ore_report.items()):
                if key == 'figures':
                    continue
                column += 2
                ws.cell(row+3, column, value)
        ws.cell(14,1,'Compressed')
        for column, value in enumerate(uncompressed_metrics):
            if value == 'figures':
                continue
            column += 2
            ws.cell(14, column, value)
        for row, ore_name in enumerate(compressed_types):
            ws.cell(row+15, 1, ore_name)
            ore_report = ice_data[ore_name]
            for column, (key, value) in enumerate(ore_report.items()):
                if key == 'figures':
                    continue
                column += 2
                ws.cell(row+15, column, value)
        ore_report_fname = 'Sleeper_iceReport_daily_'+str(today_timestamp)+'.xlsx'
        ore_report_fname = os.path.join(report_dir, ore_report_fname)
        wb.save(ore_report_fname)